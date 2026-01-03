import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
#from PIL import Image, ImageTk
import csv
import os
import json
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ==================== CONFIGURATION ====================
GAME_PRICES = {
    "SNOOKER": {
        "15min": 50, "30min": 100, "40min": 120, "45min": 150, "50min": 180,
        "hour": 200, "1hour 10min": 220, "1hour 15min": 250, "1hour 20min": 280,
        "1hour 30min": 300, "1hour 40min": 320, "1hour 50min": 380
    },
    "BLUE BOARD": {
        "15min": 40, "30min": 80, "40min": 100, "45min": 120, "50min": 140,
        "hour": 150, "1hour 10min": 170, "1hour 15min": 190, "1hour 20min": 200,
        "1hour 30min": 230, "1hour 40min": 270, "1hour 50min": 290
    },
    "BLACK BOARD": {
        "15min": 40, "30min": 80, "40min": 100, "45min": 120, "50min": 140,
        "hour": 150, "1hour 10min": 170, "1hour 15min": 190, "1hour 20min": 200,
        "1hour 30min": 230, "1hour 40min": 270, "1hour 50min": 290
    },
    "FIFA 25": {
        "15min": 40, "30min": 80, "40min": 100, "45min": 120, "50min": 140,
        "hour": 150, "1hour 10min": 170, "1hour 15min": 190, "1hour 20min": 200,
        "1hour 30min": 230, "1hour 40min": 270, "1hour 50min": 290
    },
    "GTA 5": {
        "15min": 30, "30min": 60, "40min": 70, "45min": 80, "50min": 90,
        "hour": 100, "1hour 10min": 120, "1hour 15min": 130, "1hour 20min": 140,
        "1hour 30min": 160, "1hour 40min": 170, "1hour 50min": 190
    },
    "EFOOTBALL": {
        "15min": 30, "30min": 60, "40min": 70, "45min": 80, "50min": 100,
        "hour": 120, "1hour 10min": 140, "1hour 15min": 150, "1hour 20min": 160,
        "1hour 30min": 180, "1hour 40min": 190, "1hour 50min": 220
    },
    "CALL OF DUTY": {
        "15min": 40, "30min": 80, "40min": 100, "45min": 120, "50min": 140,
        "hour": 150, "1hour 10min": 170, "1hour 15min": 190, "1hour 20min": 200,
        "1hour 30min": 230, "1hour 40min": 270, "1hour 50min": 290
    },
    "MINI BOARD": {
        "15min": 20, "30min": 50, "40min": 60, "45min": 80, "50min": 90,
        "hour": 100, "1hour 10min": 120, "1hour 15min": 130, "1hour 20min": 140,
        "1hour 30min": 150, "1hour 40min": 160, "1hour 50min": 190
    }
}

GAMES_CONFIG_FILE = "games_config.json"

# Global variables for edit mode
edit_mode = False
edit_item_id = None

# ==================== UTILITY FUNCTIONS ====================
def format_duration(hours):
    """Format duration in hours to readable format"""
    total_minutes = int(hours * 60)
    h = total_minutes // 60
    m = total_minutes % 60
    if h > 0 and m > 0:
        return f"{h}h {m}m"
    elif h > 0:
        return f"{h}h"
    else:
        return f"{m}m"

def calculate_amount(game, start, end, controllers=2):
    """Calculate amount based on game, duration, and controllers"""
    if game not in GAME_PRICES:
        return 0, 0
    
    fmt = "%I:%M %p"
    try:
        start_dt = datetime.strptime(start.strip(), fmt)
        end_dt = datetime.strptime(end.strip(), fmt)
    except Exception:
        return 0, 0
    
    if end_dt < start_dt:
        end_dt += timedelta(days=1)
    
    duration_minutes = (end_dt - start_dt).total_seconds() / 60.0
    if duration_minutes <= 0:
        return 0, 0
    
    duration_hours = duration_minutes / 60.0
    price_table = GAME_PRICES[game]
    base_price = 0
    
    if duration_minutes <= 15:
        base_price = price_table["15min"]
    elif duration_minutes <= 30:
        base_price = price_table["30min"]
    elif duration_minutes <= 40:
        base_price = price_table["40min"]
    elif duration_minutes <= 45:
        base_price = price_table["45min"]
    elif duration_minutes <= 50:
        base_price = price_table["50min"]
    elif duration_minutes <= 60:
        base_price = price_table["hour"]
    elif duration_minutes <= 70:
        base_price = price_table["1hour 10min"]
    elif duration_minutes <= 75:
        base_price = price_table["1hour 15min"]
    elif duration_minutes <= 80:
        base_price = price_table.get("1hour 20min", price_table["1hour 15min"])
    elif duration_minutes <= 90:
        base_price = price_table["1hour 30min"]
    elif duration_minutes <= 100:
        base_price = price_table["1hour 40min"]
    elif duration_minutes <= 110:
        base_price = price_table["1hour 50min"]
    else:
        full_hours = int(duration_minutes // 60)
        remaining_minutes = duration_minutes % 60
        base_price = full_hours * price_table["hour"]
        
        if remaining_minutes > 0:
            if remaining_minutes <= 15:
                base_price += price_table["15min"]
            elif remaining_minutes <= 30:
                base_price += price_table["30min"]
            elif remaining_minutes <= 40:
                base_price += price_table["40min"]
            elif remaining_minutes <= 45:
                base_price += price_table["45min"]
            elif remaining_minutes <= 50:
                base_price += price_table["50min"]
            else:
                base_price += price_table["hour"]
    
    try:
        controllers = int(controllers)
    except Exception:
        controllers = 2
    
    if game in ["FIFA 25", "EFOOTBALL"] and controllers > 2:
        base_price += (controllers - 2) * 50
    
    return base_price, duration_hours

# ==================== FILE HANDLING ====================
def get_today_folder():
    """Get or create today's folder"""
    today = datetime.now().strftime("%Y-%m-%d")
    folder_path = os.path.join("records", today)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path

def get_today_csv_filename():
    """Get today's CSV filename"""
    today = datetime.now().strftime("%Y-%m-%d")
    folder_path = get_today_folder()
    return os.path.join(folder_path, f"{today}.csv")

def get_recovery_csv_filename():
    """Get recovery CSV filename for current sessions"""
    today = datetime.now().strftime("%Y-%m-%d")
    folder_path = get_today_folder()
    return os.path.join(folder_path, f"{today}_current_sessions.csv")

def get_csv_filename_for_date(date_str):
    """Get CSV filename for a specific date"""
    folder_path = os.path.join("records", date_str)
    return os.path.join(folder_path, f"{date_str}.csv")

def get_expenses_csv_filename(date_str=None):
    """Get expenses CSV filename for a specific date"""
    if date_str is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
    folder_path = os.path.join("records", date_str)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return os.path.join(folder_path, f"{date_str}_expenses.csv")

def update_csv_record(filename, date, old_start_time, old_game, new_start_time, new_game,
                     end_time, duration, total, paid, balance, cash, gpay, controllers,
                     payment_status="PAID", customer_name=""):
    """Update a record in the CSV file"""
    records = []
    fieldnames = ["Date", "Customer Name", "Start Time", "Game", "End Time", "Duration",
                 "Total Amount", "Paid Amount", "Balance Amount", "Cash", "GPay",
                 "Controllers", "Payment Status"]
    
    if os.path.isfile(filename):
        with open(filename, "r", newline="", encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if (row.get("Date") == date and row.get("Start Time") == old_start_time and
                    row.get("Game") == old_game and row.get("End Time") != ""):
                    row["Customer Name"] = customer_name if payment_status == "PENDING" else ""
                    row["Start Time"] = new_start_time
                    row["Game"] = new_game
                    row["End Time"] = end_time
                    row["Duration"] = duration
                    row["Total Amount"] = total
                    row["Paid Amount"] = paid
                    row["Balance Amount"] = balance
                    row["Cash"] = cash
                    row["GPay"] = gpay
                    row["Controllers"] = controllers
                    row["Payment Status"] = payment_status
                
                updated_row = {field: row.get(field, "") for field in fieldnames}
                records.append(updated_row)
    
    with open(filename, "w", newline="", encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

def delete_csv_record(filename, date, start_time, game):
    """Delete a record from the CSV file"""
    if not os.path.isfile(filename):
        return
    
    records = []
    fieldnames = ["Date", "Customer Name", "Start Time", "Game", "End Time", "Duration",
                 "Total Amount", "Paid Amount", "Balance Amount", "Cash", "GPay",
                 "Controllers", "Payment Status"]
    
    with open(filename, "r", newline="", encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not (row.get("Date") == date and row.get("Start Time") == start_time and
                   row.get("Game") == game):
                updated_row = {field: row.get(field, "") for field in fieldnames}
                records.append(updated_row)
    
    with open(filename, "w", newline="", encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

# ==================== GAME CONFIGURATION MANAGEMENT ====================
def load_game_prices():
    """Load game prices from configuration file"""
    global GAME_PRICES
    if os.path.exists(GAMES_CONFIG_FILE):
        try:
            with open(GAMES_CONFIG_FILE, 'r', encoding='utf-8') as f:
                GAME_PRICES = json.load(f)
        except Exception as e:
            print(f"Error loading game prices: {e}")

def save_game_prices():
    """Save game prices to configuration file"""
    try:
        with open(GAMES_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(GAME_PRICES, f, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save game prices: {e}")
        return False

def update_game_dropdown():
    """Update the game dropdown in the main window"""
    combo_game['values'] = sorted(list(GAME_PRICES.keys()))

def import_games_from_file():
    """Import games and prices from CSV or Excel file"""
    file_path = filedialog.askopenfilename(
        title="Select Games Price File",
        filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    
    if not file_path:
        return
    
    try:
        imported_games = {}
        
        if file_path.lower().endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    game_name = row.get("Game Name", "").strip().upper()
                    if not game_name:
                        continue
                    
                    imported_games[game_name] = {
                        "15min": float(row.get("15min", 0) or 0),
                        "30min": float(row.get("30min", 0) or 0),
                        "40min": float(row.get("40min", 0) or 0),
                        "45min": float(row.get("45min", 0) or 0),
                        "50min": float(row.get("50min", 0) or 0),
                        "hour": float(row.get("hour", 0) or 0),
                        "1hour 10min": float(row.get("1hour 10min", 0) or 0),
                        "1hour 15min": float(row.get("1hour 15min", 0) or 0),
                        "1hour 20min": float(row.get("1hour 20min", 0) or 0),
                        "1hour 30min": float(row.get("1hour 30min", 0) or 0),
                        "1hour 40min": float(row.get("1hour 40min", 0) or 0),
                        "1hour 50min": float(row.get("1hour 50min", 0) or 0)
                    }
        
        elif file_path.lower().endswith('.xlsx'):
            if not EXCEL_AVAILABLE:
                messagebox.showerror("Error", 
                    "openpyxl library is required to import Excel files.\n"
                    "Install it using: pip install openpyxl")
                return
            
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                game_name = str(row[0]).strip().upper()
                
                imported_games[game_name] = {
                    "15min": float(row[1] if row[1] else 0),
                    "30min": float(row[2] if row[2] else 0),
                    "40min": float(row[3] if row[3] else 0),
                    "45min": float(row[4] if row[4] else 0),
                    "50min": float(row[5] if row[5] else 0),
                    "hour": float(row[6] if row[6] else 0),
                    "1hour 10min": float(row[7] if row[7] else 0),
                    "1hour 15min": float(row[8] if row[8] else 0),
                    "1hour 20min": float(row[9] if row[9] else 0),
                    "1hour 30min": float(row[10] if row[10] else 0),
                    "1hour 40min": float(row[11] if row[11] else 0),
                    "1hour 50min": float(row[12] if row[12] else 0)
                }
        
        if not imported_games:
            messagebox.showwarning("Warning", "No valid games found in the file")
            return
        
        preview_text = f"Found {len(imported_games)} game(s):\n\n"
        for game in list(imported_games.keys())[:10]:
            preview_text += f"• {game}\n"
        
        if len(imported_games) > 10:
            preview_text += f"... and {len(imported_games) - 10} more\n"
        
        preview_text += f"\nDo you want to import these games?"
        
        response = messagebox.askyesno("Confirm Import", preview_text)
        if not response:
            return
        
        new_count = 0
        updated_count = 0
        
        for game_name, prices in imported_games.items():
            if game_name in GAME_PRICES:
                updated_count += 1
            else:
                new_count += 1
            GAME_PRICES[game_name] = prices
        
        if save_game_prices():
            messagebox.showinfo("Import Success", 
                f"Import completed successfully!\n\n"
                f"New games: {new_count}\n"
                f"Updated games: {updated_count}\n"
                f"Total games: {len(GAME_PRICES)}")
            update_game_dropdown()
        
    except Exception as e:
        messagebox.showerror("Import Error", f"Failed to import games:\n{str(e)}")

def export_games_template():
    """Export a template file for bulk game import"""
    file_path = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="games_price_template.csv"
    )
    
    if not file_path:
        return
    
    try:
        if file_path.lower().endswith('.csv'):
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "Game Name", "15min", "30min", "40min", "45min", "50min", 
                    "hour", "1hour 10min", "1hour 15min", "1hour 20min", 
                    "1hour 30min", "1hour 40min", "1hour 50min"
                ])
                writer.writerow([
                    "SAMPLE GAME 1", 40, 80, 100, 120, 140, 150, 170, 190, 200, 230, 270, 290
                ])
                writer.writerow([
                    "SAMPLE GAME 2", 50, 100, 120, 150, 180, 200, 220, 250, 280, 300, 320, 380
                ])
            
            messagebox.showinfo("Success", 
                f"Template exported successfully!\n\n"
                f"File: {file_path}\n\n"
                f"Edit this file with your games and prices, then use "
                f"'Import Games' to load them.")
        
        elif file_path.lower().endswith('.xlsx'):
            if not EXCEL_AVAILABLE:
                messagebox.showerror("Error", 
                    "openpyxl library is required to export Excel files.\n"
                    "Install it using: pip install openpyxl")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Games Price Template"
            
            headers = [
                "Game Name", "15min", "30min", "40min", "45min", "50min", 
                "hour", "1hour 10min", "1hour 15min", "1hour 20min", 
                "1hour 30min", "1hour 40min", "1hour 50min"
            ]
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            ws.append(["SAMPLE GAME 1", 40, 80, 100, 120, 140, 150, 170, 190, 200, 230, 270, 290])
            ws.append(["SAMPLE GAME 2", 50, 100, 120, 150, 180, 200, 220, 250, 280, 300, 320, 380])
            
            ws.column_dimensions['A'].width = 20
            for col in range(2, 14):
                ws.column_dimensions[chr(64 + col)].width = 12
            
            wb.save(file_path)
            
            messagebox.showinfo("Success", 
                f"Template exported successfully!\n\n"
                f"File: {file_path}\n\n"
                f"Edit this file with your games and prices, then use "
                f"'Import Games' to load them.")
    
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export template:\n{str(e)}")

def export_current_games():
    """Export current games and prices to a file"""
    if not GAME_PRICES:
        messagebox.showinfo("Info", "No games to export")
        return
    
    file_path = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=f"current_games_{datetime.now().strftime('%Y%m%d')}.csv"
    )
    
    if not file_path:
        return
    
    try:
        if file_path.lower().endswith('.csv'):
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "Game Name", "15min", "30min", "40min", "45min", "50min", 
                    "hour", "1hour 10min", "1hour 15min", "1hour 20min", 
                    "1hour 30min", "1hour 40min", "1hour 50min"
                ])
                
                for game_name in sorted(GAME_PRICES.keys()):
                    prices = GAME_PRICES[game_name]
                    writer.writerow([
                        game_name,
                        prices.get("15min", 0),
                        prices.get("30min", 0),
                        prices.get("40min", 0),
                        prices.get("45min", 0),
                        prices.get("50min", 0),
                        prices.get("hour", 0),
                        prices.get("1hour 10min", 0),
                        prices.get("1hour 15min", 0),
                        prices.get("1hour 20min", 0),
                        prices.get("1hour 30min", 0),
                        prices.get("1hour 40min", 0),
                        prices.get("1hour 50min", 0)
                    ])
            
            messagebox.showinfo("Success", 
                f"Exported {len(GAME_PRICES)} games successfully!\n\nFile: {file_path}")
        
        elif file_path.lower().endswith('.xlsx'):
            if not EXCEL_AVAILABLE:
                messagebox.showerror("Error", 
                    "openpyxl library is required to export Excel files.\n"
                    "Install it using: pip install openpyxl")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Current Games & Prices"
            
            headers = [
                "Game Name", "15min", "30min", "40min", "45min", "50min", 
                "hour", "1hour 10min", "1hour 15min", "1hour 20min", 
                "1hour 30min", "1hour 40min", "1hour 50min"
            ]
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            for game_name in sorted(GAME_PRICES.keys()):
                prices = GAME_PRICES[game_name]
                ws.append([
                    game_name,
                    prices.get("15min", 0),
                    prices.get("30min", 0),
                    prices.get("40min", 0),
                    prices.get("45min", 0),
                    prices.get("50min", 0),
                    prices.get("hour", 0),
                    prices.get("1hour 10min", 0),
                    prices.get("1hour 15min", 0),
                    prices.get("1hour 20min", 0),
                    prices.get("1hour 30min", 0),
                    prices.get("1hour 40min", 0),
                    prices.get("1hour 50min", 0)
                ])
            
            ws.column_dimensions['A'].width = 20
            for col in range(2, 14):
                ws.column_dimensions[chr(64 + col)].width = 12
            
            wb.save(file_path)
            
            messagebox.showinfo("Success", 
                f"Exported {len(GAME_PRICES)} games successfully!\n\nFile: {file_path}")
    
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export games:\n{str(e)}")

def manage_games():
    """Window to manage games - add, edit, delete, import, export"""
    games_window = tk.Toplevel(root)
    games_window.title("Manage Games & Prices")
    games_window.geometry("900x650")
    games_window.transient(root)
    games_window.grab_set()
    
    title_label = tk.Label(games_window, text="Game Management", 
                          font=("Arial", 16, "bold"), bg="#2C3E50", fg="white")
    title_label.pack(fill="x", pady=(0, 10))
    
    list_frame = tk.Frame(games_window)
    list_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    list_scroll = tk.Scrollbar(list_frame)
    list_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    games_listbox = tk.Listbox(list_frame, font=("Arial", 12), 
                               yscrollcommand=list_scroll.set, height=15)
    games_listbox.pack(side=tk.LEFT, fill="both", expand=True)
    list_scroll.config(command=games_listbox.yview)
    
    def refresh_games_list():
        games_listbox.delete(0, tk.END)
        for game_name in sorted(GAME_PRICES.keys()):
            games_listbox.insert(tk.END, game_name)
    
    def add_new_game():
        """Add a new game with pricing"""
        add_window = tk.Toplevel(games_window)
        add_window.title("Add New Game")
        add_window.geometry("500x650")
        add_window.transient(games_window)
        add_window.grab_set()
        
        tk.Label(add_window, text="Add New Game", font=("Arial", 14, "bold")).pack(pady=10)
        
        tk.Label(add_window, text="Game Name:", font=("Arial", 11)).pack(pady=(10, 5))
        game_name_entry = tk.Entry(add_window, width=40, font=("Arial", 11))
        game_name_entry.pack(pady=5)
        
        prices_frame = tk.Frame(add_window)
        prices_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        tk.Label(prices_frame, text="Set Prices for Different Durations", 
                font=("Arial", 11, "bold")).grid(row=0, column=0, columnspan=2, pady=10)
        
        durations = [
            ("15min", "15 Minutes"),
            ("30min", "30 Minutes"),
            ("40min", "40 Minutes"),
            ("45min", "45 Minutes"),
            ("50min", "50 Minutes"),
            ("hour", "1 Hour"),
            ("1hour 10min", "1 Hour 10 Minutes"),
            ("1hour 15min", "1 Hour 15 Minutes"),
            ("1hour 20min", "1 Hour 20 Minutes"),
            ("1hour 30min", "1 Hour 30 Minutes"),
            ("1hour 40min", "1 Hour 40 Minutes"),
            ("1hour 50min", "1 Hour 50 Minutes")
        ]
        
        price_entries = {}
        
        for idx, (key, label) in enumerate(durations, start=1):
            tk.Label(prices_frame, text=f"{label}:", 
                    font=("Arial", 10)).grid(row=idx, column=0, sticky="e", padx=5, pady=3)
            entry = tk.Entry(prices_frame, width=15, font=("Arial", 10))
            entry.grid(row=idx, column=1, sticky="w", padx=5, pady=3)
            entry.insert(0, "0")
            price_entries[key] = entry
        
        def save_new_game():
            game_name = game_name_entry.get().strip().upper()
            
            if not game_name:
                messagebox.showerror("Error", "Please enter a game name", parent=add_window)
                return
            
            if game_name in GAME_PRICES:
                messagebox.showerror("Error", "This game already exists!", parent=add_window)
                return
            
            new_prices = {}
            try:
                for key, entry in price_entries.items():
                    price = float(entry.get().strip())
                    if price < 0:
                        raise ValueError("Negative price")
                    new_prices[key] = price
            except ValueError:
                messagebox.showerror("Error", "Please enter valid positive numbers for all prices", 
                                   parent=add_window)
                return
            
            GAME_PRICES[game_name] = new_prices
            
            if save_game_prices():
                messagebox.showinfo("Success", f"Game '{game_name}' added successfully!", 
                                  parent=add_window)
                refresh_games_list()
                update_game_dropdown()
                add_window.destroy()
        
        btn_frame = tk.Frame(add_window)
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Save Game", command=save_new_game, 
                 bg="green", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=add_window.destroy, 
                 bg="gray", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    
    def edit_selected_game():
        """Edit the selected game's prices"""
        selection = games_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a game to edit", parent=games_window)
            return
        
        game_name = games_listbox.get(selection[0])
        
        edit_window = tk.Toplevel(games_window)
        edit_window.title(f"Edit Game: {game_name}")
        edit_window.geometry("500x650")
        edit_window.transient(games_window)
        edit_window.grab_set()
        
        tk.Label(edit_window, text=f"Edit Game: {game_name}", 
                font=("Arial", 14, "bold")).pack(pady=10)
        
        prices_frame = tk.Frame(edit_window)
        prices_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        tk.Label(prices_frame, text="Update Prices for Different Durations", 
                font=("Arial", 11, "bold")).grid(row=0, column=0, columnspan=2, pady=10)
        
        durations = [
            ("15min", "15 Minutes"),
            ("30min", "30 Minutes"),
            ("40min", "40 Minutes"),
            ("45min", "45 Minutes"),
            ("50min", "50 Minutes"),
            ("hour", "1 Hour"),
            ("1hour 10min", "1 Hour 10 Minutes"),
            ("1hour 15min", "1 Hour 15 Minutes"),
            ("1hour 20min", "1 Hour 20 Minutes"),
            ("1hour 30min", "1 Hour 30 Minutes"),
            ("1hour 40min", "1 Hour 40 Minutes"),
            ("1hour 50min", "1 Hour 50 Minutes")
        ]
        
        price_entries = {}
        current_prices = GAME_PRICES[game_name]
        
        for idx, (key, label) in enumerate(durations, start=1):
            tk.Label(prices_frame, text=f"{label}:", 
                    font=("Arial", 10)).grid(row=idx, column=0, sticky="e", padx=5, pady=3)
            entry = tk.Entry(prices_frame, width=15, font=("Arial", 10))
            entry.grid(row=idx, column=1, sticky="w", padx=5, pady=3)
            entry.insert(0, str(current_prices.get(key, 0)))
            price_entries[key] = entry
        
        def save_edited_game():
            updated_prices = {}
            try:
                for key, entry in price_entries.items():
                    price = float(entry.get().strip())
                    if price < 0:
                        raise ValueError("Negative price")
                    updated_prices[key] = price
            except ValueError:
                messagebox.showerror("Error", "Please enter valid positive numbers for all prices", 
                                   parent=edit_window)
                return
            
            GAME_PRICES[game_name] = updated_prices
            
            if save_game_prices():
                messagebox.showinfo("Success", f"Game '{game_name}' updated successfully!", 
                                  parent=edit_window)
                refresh_games_list()
                edit_window.destroy()
        
        btn_frame = tk.Frame(edit_window)
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Save Changes", command=save_edited_game, 
                 bg="green", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=edit_window.destroy, 
                 bg="gray", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    
    def delete_selected_game():
        """Delete the selected game"""
        selection = games_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a game to delete", parent=games_window)
            return
        
        game_name = games_listbox.get(selection[0])
        
        response = messagebox.askyesno("Confirm Delete", 
                                      f"Are you sure you want to delete '{game_name}'?\n\n"
                                      f"This action cannot be undone!", 
                                      parent=games_window)
        if not response:
            return
        
        del GAME_PRICES[game_name]
        
        if save_game_prices():
            messagebox.showinfo("Success", f"Game '{game_name}' deleted successfully!", 
                              parent=games_window)
            refresh_games_list()
            update_game_dropdown()
    
    def view_game_details():
        """View detailed pricing for selected game"""
        selection = games_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a game to view", parent=games_window)
            return
        
        game_name = games_listbox.get(selection[0])
        prices = GAME_PRICES[game_name]
        
        details = f"Pricing for {game_name}\n" + "="*40 + "\n\n"
        
        duration_labels = {
            "15min": "15 Minutes",
            "30min": "30 Minutes",
            "40min": "40 Minutes",
            "45min": "45 Minutes",
            "50min": "50 Minutes",
            "hour": "1 Hour",
            "1hour 10min": "1 Hour 10 Minutes",
            "1hour 15min": "1 Hour 15 Minutes",
            "1hour 20min": "1 Hour 20 Minutes",
            "1hour 30min": "1 Hour 30 Minutes",
            "1hour 40min": "1 Hour 40 Minutes",
            "1hour 50min": "1 Hour 50 Minutes"
        }
        
        for key in ["15min", "30min", "40min", "45min", "50min", "hour", 
                   "1hour 10min", "1hour 15min", "1hour 20min", 
                   "1hour 30min", "1hour 40min", "1hour 50min"]:
            label = duration_labels.get(key, key)
            price = prices.get(key, 0)
            details += f"{label:25} Rs. {price}\n"
        
        messagebox.showinfo(f"Game Details - {game_name}", details, parent=games_window)
    
    button_frame = tk.Frame(games_window)
    button_frame.pack(fill="x", padx=10, pady=5)
    
    tk.Button(button_frame, text="Add New Game", command=add_new_game, 
             bg="green", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Edit Selected", command=edit_selected_game, 
             bg="orange", fg="black", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Delete Selected", command=delete_selected_game, 
             bg="darkred", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="View Details", command=view_game_details, 
             bg="blue", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    
    import_export_frame = tk.Frame(games_window)
    import_export_frame.pack(fill="x", padx=10, pady=5)
    
    tk.Button(import_export_frame, text="📥 Import Games", command=import_games_from_file, 
             bg="#28a745", fg="white", width=15, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
    tk.Button(import_export_frame, text="📤 Export Current", command=export_current_games, 
             bg="#17a2b8", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(import_export_frame, text="📋 Export Template", command=export_games_template, 
             bg="#6c757d", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(import_export_frame, text="Close", command=games_window.destroy, 
             bg="gray", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    
    refresh_games_list()

# ==================== EXPENSE MANAGEMENT ====================
def add_expense():
    """Add a new expense entry"""
    expense_window = tk.Toplevel(root)
    expense_window.title("Add Expense")
    expense_window.geometry("450x350")
    expense_window.transient(root)
    expense_window.grab_set()
    
    tk.Label(expense_window, text="Date (YYYY-MM-DD):", font=("Arial", 10)).pack(pady=(15, 5))
    date_entry = tk.Entry(expense_window, width=30, font=("Arial", 10))
    date_entry.pack(pady=5)
    date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
    
    tk.Label(expense_window, text="Category:", font=("Arial", 10)).pack(pady=(10, 5))
    category_var = tk.StringVar()
    category_combo = ttk.Combobox(expense_window, textvariable=category_var,
                                 values=["Electricity", "Rent", "Maintenance", "Equipment",
                                        "Staff Salary", "Food & Beverages", "Internet", "Other"],
                                 width=28, font=("Arial", 10))
    category_combo.pack(pady=5)
    
    tk.Label(expense_window, text="Description:", font=("Arial", 10)).pack(pady=(10, 5))
    desc_entry = tk.Entry(expense_window, width=30, font=("Arial", 10))
    desc_entry.pack(pady=5)
    
    tk.Label(expense_window, text="Amount (Rs):", font=("Arial", 10)).pack(pady=(10, 5))
    amount_entry = tk.Entry(expense_window, width=30, font=("Arial", 10))
    amount_entry.pack(pady=5)
    
    def save_expense():
        date = date_entry.get().strip()
        category = category_var.get().strip()
        description = desc_entry.get().strip()
        amount = amount_entry.get().strip()
        
        if not date or not category or not amount:
            messagebox.showerror("Error", "Please fill Date, Category, and Amount",
                               parent=expense_window)
            return
        
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD",
                               parent=expense_window)
            return
        
        try:
            float(amount)
        except ValueError:
            messagebox.showerror("Error", "Invalid amount", parent=expense_window)
            return
        
        filename = get_expenses_csv_filename(date)
        file_exists = os.path.isfile(filename)
        
        try:
            with open(filename, "a", newline="", encoding='utf-8') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow(["Date", "Category", "Description", "Amount"])
                writer.writerow([date, category, description, amount])
            
            messagebox.showinfo("Success", "Expense added successfully!", parent=expense_window)
            expense_window.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save expense: {e}", parent=expense_window)
    
    btn_frame = tk.Frame(expense_window)
    btn_frame.pack(pady=20)
    
    tk.Button(btn_frame, text="Save Expense", command=save_expense,
             bg="green", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="Cancel", command=expense_window.destroy,
             bg="gray", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)

def view_expenses():
    """View all expenses with filters"""
    expenses_window = tk.Toplevel(root)
    expenses_window.title("View Expenses")
    expenses_window.geometry("1000x600")
    
    filter_frame = tk.Frame(expenses_window, bg="#e7f3ff", relief=tk.RIDGE, bd=2)
    filter_frame.pack(fill="x", padx=10, pady=10)
    
    tk.Label(filter_frame, text="Date Range:", font=("Arial", 11, "bold"),
            bg="#e7f3ff").pack(side=tk.LEFT, padx=10, pady=8)
    
    tk.Label(filter_frame, text="From:", bg="#e7f3ff").pack(side=tk.LEFT, padx=5)
    from_date_entry = tk.Entry(filter_frame, width=12, font=("Arial", 10))
    from_date_entry.pack(side=tk.LEFT, padx=5)
    from_date_entry.insert(0, (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
    
    tk.Label(filter_frame, text="To:", bg="#e7f3ff").pack(side=tk.LEFT, padx=5)
    to_date_entry = tk.Entry(filter_frame, width=12, font=("Arial", 10))
    to_date_entry.pack(side=tk.LEFT, padx=5)
    to_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
    
    tree_frame = tk.Frame(expenses_window)
    tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    expenses_tree = ttk.Treeview(tree_frame,
                                columns=("Date", "Category", "Description", "Amount"),
                                show="headings", yscrollcommand=tree_scroll.set)
    tree_scroll.config(command=expenses_tree.yview)
    
    expenses_tree.heading("Date", text="Date")
    expenses_tree.heading("Category", text="Category")
    expenses_tree.heading("Description", text="Description")
    expenses_tree.heading("Amount", text="Amount (Rs)")
    
    expenses_tree.column("Date", width=120)
    expenses_tree.column("Category", width=150)
    expenses_tree.column("Description", width=400)
    expenses_tree.column("Amount", width=120, anchor=tk.E)
    
    expenses_tree.pack(fill="both", expand=True)
    
    summary_frame = tk.Frame(expenses_window, bg="#fff3cd", relief=tk.RIDGE, bd=2)
    summary_frame.pack(fill="x", padx=10, pady=10)
    
    summary_label = tk.Label(summary_frame, text="", font=("Arial", 11, "bold"), bg="#fff3cd")
    summary_label.pack(pady=10)
    
    def load_expenses():
        for item in expenses_tree.get_children():
            expenses_tree.delete(item)
        
        try:
            from_date = datetime.strptime(from_date_entry.get().strip(), "%Y-%m-%d")
            to_date = datetime.strptime(to_date_entry.get().strip(), "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD",
                               parent=expenses_window)
            return
        
        if from_date > to_date:
            messagebox.showerror("Error", "From date must be before To date",
                               parent=expenses_window)
            return
        
        records_folder = "records"
        total_expenses = 0.0
        expense_count = 0
        
        if os.path.exists(records_folder):
            for folder_name in sorted(os.listdir(records_folder)):
                folder_path = os.path.join(records_folder, folder_name)
                if not os.path.isdir(folder_path):
                    continue
                
                try:
                    folder_date = datetime.strptime(folder_name, "%Y-%m-%d")
                    if from_date <= folder_date <= to_date:
                        expenses_file = os.path.join(folder_path, f"{folder_name}_expenses.csv")
                        if os.path.exists(expenses_file):
                            with open(expenses_file, "r", newline="", encoding='utf-8') as f:
                                reader = csv.DictReader(f)
                                for row in reader:
                                    date = row.get("Date", "")
                                    category = row.get("Category", "")
                                    description = row.get("Description", "")
                                    amount = row.get("Amount", "0")
                                    
                                    expenses_tree.insert("", "end",
                                                       values=(date, category, description,
                                                              f"Rs. {amount}"))
                                    try:
                                        total_expenses += float(amount)
                                        expense_count += 1
                                    except ValueError:
                                        pass
                except ValueError:
                    continue
        
        summary_label.config(text=f"Total Expenses: {expense_count} entries | "
                                 f"Total Amount: Rs. {total_expenses:.2f}")
    
    def delete_selected_expense():
        selected = expenses_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an expense to delete",
                                 parent=expenses_window)
            return
        
        values = expenses_tree.item(selected[0], "values")
        date = values[0]
        category = values[1]
        description = values[2]
        
        response = messagebox.askyesno("Confirm Delete",
                                      f"Delete this expense?\n\nDate: {date}\n"
                                      f"Category: {category}\nDescription: {description}",
                                      parent=expenses_window)
        if not response:
            return
        
        expenses_file = get_expenses_csv_filename(date)
        if not os.path.exists(expenses_file):
            messagebox.showerror("Error", "Expense file not found", parent=expenses_window)
            return
        
        try:
            records = []
            with open(expenses_file, "r", newline="", encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if not (row.get("Date") == date and row.get("Category") == category and
                           row.get("Description") == description):
                        records.append(row)
            
            with open(expenses_file, "w", newline="", encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=["Date", "Category", "Description", "Amount"])
                writer.writeheader()
                writer.writerows(records)
            
            messagebox.showinfo("Success", "Expense deleted successfully!", parent=expenses_window)
            load_expenses()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete expense: {e}", parent=expenses_window)
    
    button_frame = tk.Frame(expenses_window)
    button_frame.pack(fill="x", padx=10, pady=5)
    
    tk.Button(button_frame, text="Load Expenses", command=load_expenses,
             bg="blue", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Delete Selected", command=delete_selected_expense,
             bg="darkred", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Close", command=expenses_window.destroy,
             bg="gray", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    
    load_expenses()

# ==================== REPORTING FUNCTIONS ====================
def show_daily_income():
    today = datetime.now().strftime("%Y-%m-%d")
    filename = get_today_csv_filename()
    
    if not os.path.exists(filename):
        messagebox.showinfo("Daily Income", f"No finalized records found for {today}.")
        return
    
    total_income = 0.0
    cash_total = 0.0
    gpay_total = 0.0
    pending_total = 0.0
    
    try:
        with open(filename, "r", newline="", encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    amount = float(row.get("Total Amount", 0) or 0)
                    payment_status = row.get("Payment Status", "PAID")
                    
                    if payment_status == "PENDING":
                        pending_total += amount
                    else:
                        total_income += amount
                        cash_total += float(row.get("Cash", 0) or 0)
                        gpay_total += float(row.get("GPay", 0) or 0)
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read records: {e}")
        return
    
    msg = f"Income for {today}:\n\n"
    msg += f"Total (PAID): Rs. {total_income:.2f}\n"
    msg += f"Cash: Rs. {cash_total:.2f}\n"
    msg += f"GPay: Rs. {gpay_total:.2f}\n"
    
    if pending_total > 0:
        msg += f"\n⚠ Pending Payments: Rs. {pending_total:.2f}"
    
    messagebox.showinfo("Daily Income", msg)

def show_monthly_income():
    today = datetime.now()
    current_month = today.month
    current_year = today.year
    
    records_folder = "records"
    if not os.path.exists(records_folder):
        messagebox.showinfo("Monthly Income", "No records folder found.")
        return
    
    total_income = 0.0
    cash_total = 0.0
    gpay_total = 0.0
    pending_total = 0.0
    days_count = 0
    
    for folder_name in os.listdir(records_folder):
        folder_path = os.path.join(records_folder, folder_name)
        if not os.path.isdir(folder_path):
            continue
        
        try:
            folder_date = datetime.strptime(folder_name, "%Y-%m-%d")
            if folder_date.month == current_month and folder_date.year == current_year:
                csv_filename = os.path.join(folder_path, f"{folder_name}.csv")
                if os.path.exists(csv_filename):
                    days_count += 1
                    with open(csv_filename, "r", newline="", encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            try:
                                amount = float(row.get("Total Amount", 0) or 0)
                                payment_status = row.get("Payment Status", "PAID")
                                
                                if payment_status == "PENDING":
                                    pending_total += amount
                                else:
                                    total_income += amount
                                    cash_total += float(row.get("Cash", 0) or 0)
                                    gpay_total += float(row.get("GPay", 0) or 0)
                            except (ValueError, TypeError):
                                pass
        except ValueError:
            continue
    
    if days_count == 0:
        messagebox.showinfo("Monthly Income", f"No records found for {today.strftime('%B %Y')}.")
        return
    
    month_name = today.strftime("%B %Y")
    msg = f"Total income for {month_name}:\n"
    msg += f"(Based on {days_count} day(s) of records)\n\n"
    msg += f"Total (PAID): Rs. {total_income:.2f}\n"
    msg += f"Cash: Rs. {cash_total:.2f}\n"
    msg += f"GPay: Rs. {gpay_total:.2f}\n"
    
    if pending_total > 0:
        msg += f"\n⚠ Pending Payments: Rs. {pending_total:.2f}"
    
    messagebox.showinfo("Monthly Income", msg)

def add_manual_daily_summary():
    """Add manual daily summary for missing records"""
    date_str = simpledialog.askstring("Manual Summary Entry",
                                     "Enter the DATE for the missing record (YYYY-MM-DD):",
                                     parent=root)
    if not date_str:
        return
    
    try:
        record_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        messagebox.showerror("Error", "Invalid Date format. Please use YYYY-MM-DD.")
        return
    
    try:
        total_amount = simpledialog.askfloat("Manual Summary Entry",
                                            f"Enter TOTAL AMOUNT for {record_date}:",
                                            parent=root)
        if total_amount is None:
            return
        total_amount = round(total_amount, 2)
    except Exception:
        messagebox.showerror("Error", "Invalid Total Amount.")
        return
    
    try:
        cash_amount = simpledialog.askfloat("Manual Summary Entry",
                                           f"Enter TOTAL CASH for {record_date}:",
                                           parent=root)
        if cash_amount is None:
            return
        cash_amount = round(cash_amount, 2)
    except Exception:
        messagebox.showerror("Error", "Invalid Cash Amount.")
        return
    
    try:
        gpay_amount = simpledialog.askfloat("Manual Summary Entry",
                                           f"Enter TOTAL GPAY for {record_date}:",
                                           parent=root)
        if gpay_amount is None:
            return
        gpay_amount = round(gpay_amount, 2)
    except Exception:
        messagebox.showerror("Error", "Invalid GPay Amount.")
        return
    
    paid_amount = cash_amount + gpay_amount
    balance_amount = total_amount - paid_amount
    
    if abs(balance_amount) > 0.01:
        if not messagebox.askyesno("Warning",
                                   f"Cash + GPay (Rs. {paid_amount:.2f}) does not equal "
                                   f"Total Amount (Rs. {total_amount:.2f}). "
                                   f"Balance: Rs. {balance_amount:.2f}. "
                                   f"Do you want to save it anyway?"):
            return
    
    folder_path = os.path.join("records", record_date)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    filename = os.path.join(folder_path, f"{record_date}.csv")
    file_exists = os.path.isfile(filename)
    
    record_data = [
        record_date, "", "00:00 AM", "MANUAL SUMMARY", "00:00 AM",
        "0h 0m", f"{total_amount:.2f}", f"{paid_amount:.2f}",
        f"{balance_amount:.2f}", f"{cash_amount:.2f}", f"{gpay_amount:.2f}",
        0, "PAID"
    ]
    
    try:
        with open(filename, "a", newline="", encoding='utf-8') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["Date", "Customer Name", "Start Time", "Game", "End Time",
                               "Duration", "Total Amount", "Paid Amount", "Balance Amount",
                               "Cash", "GPay", "Controllers", "Payment Status"])
            writer.writerow(record_data)
        
        messagebox.showinfo("Success", f"Manual summary added successfully for {record_date}!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save manual summary: {e}")

def export_to_excel():
    """Enhanced Excel export with income, expenses, and net profit"""
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Error",
                           "openpyxl library is not installed.\n\n"
                           "Please install it using:\npip install openpyxl")
        return
    
    records_folder = "records"
    if not os.path.exists(records_folder):
        messagebox.showinfo("Export", "No records folder found.")
        return
    
    data = []
    for folder_name in sorted(os.listdir(records_folder)):
        folder_path = os.path.join(records_folder, folder_name)
        if not os.path.isdir(folder_path):
            continue
        
        try:
            folder_date = datetime.strptime(folder_name, "%Y-%m-%d")
            
            csv_filename = os.path.join(folder_path, f"{folder_name}.csv")
            daily_total = 0.0
            daily_cash = 0.0
            daily_gpay = 0.0
            daily_pending = 0.0
            
            if os.path.exists(csv_filename):
                with open(csv_filename, "r", newline="", encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        try:
                            amount = float(row.get("Total Amount", 0) or 0)
                            payment_status = row.get("Payment Status", "PAID")
                            
                            if payment_status == "PENDING":
                                daily_pending += amount
                            else:
                                daily_total += amount
                                daily_cash += float(row.get("Cash", 0) or 0)
                                daily_gpay += float(row.get("GPay", 0) or 0)
                        except (ValueError, TypeError):
                            pass
            
            expenses_filename = os.path.join(folder_path, f"{folder_name}_expenses.csv")
            daily_expenses = 0.0
            
            if os.path.exists(expenses_filename):
                with open(expenses_filename, "r", newline="", encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        try:
                            daily_expenses += float(row.get("Amount", 0) or 0)
                        except (ValueError, TypeError):
                            pass
            
            data.append({
                "Date": folder_date.strftime("%d-%m-%Y"),
                "Total Income": daily_total,
                "Cash": daily_cash,
                "GPay": daily_gpay,
                "Pending": daily_pending,
                "Expenses": daily_expenses,
                "Net Profit": daily_total - daily_expenses
            })
        except ValueError:
            continue
    
    if not data:
        messagebox.showinfo("Export", "No data to export.")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Income & Expenses"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    total_font = Font(bold=True, size=11)
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    expense_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    profit_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    loss_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Date", "Total Income (Rs)", "Cash (Rs)", "GPay (Rs)", "Pending (Rs)",
              "Expenses (Rs)", "Net Profit (Rs)"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    total_income = 0.0
    total_cash = 0.0
    total_gpay = 0.0
    total_pending = 0.0
    total_expenses = 0.0
    total_profit = 0.0
    
    for row_num, record in enumerate(data, 2):
        ws.cell(row=row_num, column=1).value = record["Date"]
        ws.cell(row=row_num, column=2).value = record["Total Income"]
        ws.cell(row=row_num, column=3).value = record["Cash"]
        ws.cell(row=row_num, column=4).value = record["GPay"]
        ws.cell(row=row_num, column=5).value = record["Pending"]
        ws.cell(row=row_num, column=6).value = record["Expenses"]
        ws.cell(row=row_num, column=7).value = record["Net Profit"]
        
        total_income += record["Total Income"]
        total_cash += record["Cash"]
        total_gpay += record["GPay"]
        total_pending += record["Pending"]
        total_expenses += record["Expenses"]
        total_profit += record["Net Profit"]
        
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col > 1:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")
            
            if col == 5 and record["Pending"] > 0:
                cell.fill = pending_fill
            elif col == 6 and record["Expenses"] > 0:
                cell.fill = expense_fill
            elif col == 7:
                if record["Net Profit"] > 0:
                    cell.fill = profit_fill
                elif record["Net Profit"] < 0:
                    cell.fill = loss_fill
    
    total_row = len(data) + 2
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=2).value = total_income
    ws.cell(row=total_row, column=3).value = total_cash
    ws.cell(row=total_row, column=4).value = total_gpay
    ws.cell(row=total_row, column=5).value = total_pending
    ws.cell(row=total_row, column=6).value = total_expenses
    ws.cell(row=total_row, column=7).value = total_profit
    
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        if col > 1:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="right")
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=f"Gaming_Lounge_Financial_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    )
    
    if file_path:
        try:
            wb.save(file_path)
            messagebox.showinfo("Success",
                              f"Excel file exported successfully!\n\n"
                              f"Total Income: Rs. {total_income:.2f}\n"
                              f"Total Expenses: Rs. {total_expenses:.2f}\n"
                              f"Net Profit: Rs. {total_profit:.2f}\n\n"
                              f"File saved at:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file:\n{str(e)}")

# ==================== PENDING PAYMENTS & STATEMENTS ====================
def view_pending_payments():
    """Show all pending payments with customer details"""
    pending_window = tk.Toplevel(root)
    pending_window.title("Pending Payments")
    pending_window.geometry("1000x550")
    
    tree_frame = tk.Frame(pending_window)
    tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    pending_tree = ttk.Treeview(tree_frame,
                                columns=("Date", "Customer", "Game", "Start", "Total", "Paid", "Balance"),
                                show="headings", yscrollcommand=tree_scroll.set)
    tree_scroll.config(command=pending_tree.yview)
    
    pending_tree.heading("Date", text="Date")
    pending_tree.heading("Customer", text="Customer Name")
    pending_tree.heading("Game", text="Game")
    pending_tree.heading("Start", text="Start Time")
    pending_tree.heading("Total", text="Total Amount")
    pending_tree.heading("Paid", text="Paid Amount")
    pending_tree.heading("Balance", text="Balance Amount")
    
    pending_tree.column("Date", width=100)
    pending_tree.column("Customer", width=150)
    pending_tree.column("Game", width=120)
    pending_tree.column("Start", width=100)
    pending_tree.column("Total", width=110, anchor=tk.E)
    pending_tree.column("Paid", width=110, anchor=tk.E)
    pending_tree.column("Balance", width=110, anchor=tk.E)
    
    pending_tree.pack(fill="both", expand=True)
    
    records_folder = "records"
    total_pending = 0.0
    
    if os.path.exists(records_folder):
        for folder_name in sorted(os.listdir(records_folder), reverse=True):
            folder_path = os.path.join(records_folder, folder_name)
            if not os.path.isdir(folder_path):
                continue
            
            csv_filename = os.path.join(folder_path, f"{folder_name}.csv")
            if os.path.exists(csv_filename):
                try:
                    with open(csv_filename, "r", newline="", encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            payment_status = row.get("Payment Status", "PAID")
                            if payment_status == "PENDING":
                                date = row.get("Date", "")
                                customer_name = row.get("Customer Name", "Unknown")
                                game = row.get("Game", "")
                                start_time = row.get("Start Time", "")
                                total = row.get("Total Amount", "0")
                                paid = row.get("Paid Amount", "0")
                                balance = row.get("Balance Amount", "0")
                                
                                pending_tree.insert("", "end",
                                                   values=(date, customer_name, game, start_time,
                                                          f"Rs. {total}", f"Rs. {paid}",
                                                          f"Rs. {balance}"),
                                                   tags=("pending",))
                                try:
                                    total_pending += float(balance)
                                except ValueError:
                                    pass
                except Exception as e:
                    print(f"Error reading {csv_filename}: {e}")
    
    pending_tree.tag_configure("pending", background="#ffcccc")
    
    summary_frame = tk.Frame(pending_window, bg="#fff3cd", relief=tk.RIDGE, bd=2)
    summary_frame.pack(fill="x", padx=10, pady=10)
    
    summary_label = tk.Label(summary_frame,
                            text=f"Total Pending Amount: Rs. {total_pending:.2f}",
                            font=("Arial", 14, "bold"), bg="#fff3cd", fg="#856404")
    summary_label.pack(pady=10)
    
    button_frame = tk.Frame(pending_window)
    button_frame.pack(fill="x", padx=10, pady=5)
    
    def clear_selected_pending():
        selected = pending_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a pending payment to clear",
                                 parent=pending_window)
            return
        
        values = pending_tree.item(selected[0], "values")
        date = values[0]
        customer_name = values[1]
        game = values[2]
        start_time = values[3]
        balance_str = values[6].replace("Rs. ", "")
        
        try:
            balance = float(balance_str)
        except ValueError:
            messagebox.showerror("Error", "Invalid balance amount", parent=pending_window)
            return
        
        payment_dialog = tk.Toplevel(pending_window)
        payment_dialog.title("Clear Pending Payment")
        payment_dialog.geometry("400x280")
        payment_dialog.transient(pending_window)
        payment_dialog.grab_set()
        
        info_frame = tk.Frame(payment_dialog, bg="#f0f0f0", relief=tk.RIDGE, bd=2)
        info_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(info_frame, text=f"Customer: {customer_name}", font=("Arial", 10),
                justify=tk.LEFT, bg="#f0f0f0").pack(anchor="w", padx=10, pady=2)
        tk.Label(info_frame, text=f"Date: {date}", font=("Arial", 10),
                justify=tk.LEFT, bg="#f0f0f0").pack(anchor="w", padx=10, pady=2)
        tk.Label(info_frame, text=f"Game: {game}", font=("Arial", 10),
                justify=tk.LEFT, bg="#f0f0f0").pack(anchor="w", padx=10, pady=2)
        tk.Label(info_frame, text=f"Balance: Rs. {balance:.2f}", font=("Arial", 11, "bold"),
                justify=tk.LEFT, bg="#f0f0f0", fg="red").pack(anchor="w", padx=10, pady=2)
        
        tk.Label(payment_dialog, text="Cash Payment:", font=("Arial", 10)).pack(pady=(10, 0))
        cash_entry = tk.Entry(payment_dialog, width=20, font=("Arial", 10))
        cash_entry.pack(pady=5)
        cash_entry.insert(0, str(balance))
        
        tk.Label(payment_dialog, text="GPay Payment:", font=("Arial", 10)).pack()
        gpay_entry = tk.Entry(payment_dialog, width=20, font=("Arial", 10))
        gpay_entry.pack(pady=5)
        gpay_entry.insert(0, "0")
        
        def process_payment():
            try:
                cash_amt = float(cash_entry.get() or 0)
                gpay_amt = float(gpay_entry.get() or 0)
                total_payment = cash_amt + gpay_amt
            except ValueError:
                messagebox.showerror("Error", "Invalid payment amount", parent=payment_dialog)
                return
            
            if total_payment < balance - 0.01:
                messagebox.showerror("Error",
                                   f"Payment (Rs. {total_payment:.2f}) is less than balance "
                                   f"(Rs. {balance:.2f})", parent=payment_dialog)
                return
            
            csv_filename = get_csv_filename_for_date(date)
            if not os.path.exists(csv_filename):
                messagebox.showerror("Error", f"Record file not found for {date}",
                                   parent=payment_dialog)
                return
            
            records = []
            fieldnames = ["Date", "Customer Name", "Start Time", "Game", "End Time", "Duration",
                         "Total Amount", "Paid Amount", "Balance Amount", "Cash", "GPay",
                         "Controllers", "Payment Status"]
            updated = False
            
            try:
                with open(csv_filename, "r", newline="", encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if (row.get("Date") == date and row.get("Customer Name") == customer_name and
                            row.get("Game") == game and row.get("Start Time") == start_time and
                            row.get("Payment Status") == "PENDING" and not updated):
                            
                            old_cash = float(row.get("Cash", 0) or 0)
                            old_gpay = float(row.get("GPay", 0) or 0)
                            total_amt = float(row.get("Total Amount", 0) or 0)
                            
                            row["Cash"] = f"{old_cash + cash_amt:.2f}"
                            row["GPay"] = f"{old_gpay + gpay_amt:.2f}"
                            row["Paid Amount"] = f"{total_amt:.2f}"
                            row["Balance Amount"] = "0.00"
                            row["Payment Status"] = "PAID"
                            row["Customer Name"] = ""
                            updated = True
                        
                        updated_row = {field: row.get(field, "") for field in fieldnames}
                        records.append(updated_row)
                
                with open(csv_filename, "w", newline="", encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(records)
                
                if updated:
                    messagebox.showinfo("Success",
                                      f"Payment cleared successfully!\nPaid: Rs. {total_payment:.2f}",
                                      parent=payment_dialog)
                    payment_dialog.destroy()
                    pending_window.destroy()
                    view_pending_payments()
                else:
                    messagebox.showerror("Error", "Could not find the record to update",
                                       parent=payment_dialog)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update payment: {e}",
                                   parent=payment_dialog)
        
        btn_frame = tk.Frame(payment_dialog)
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="Process Payment", command=process_payment,
                 bg="green", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=payment_dialog.destroy,
                 bg="gray", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    
    def delete_selected_pending():
        selected = pending_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a pending payment to delete",
                                 parent=pending_window)
            return
        
        values = pending_tree.item(selected[0], "values")
        date = values[0]
        customer_name = values[1]
        game = values[2]
        start_time = values[3]
        
        response = messagebox.askyesno("Confirm Delete",
                                      f"Are you sure you want to delete this record?\n\n"
                                      f"Date: {date}\nCustomer: {customer_name}\nGame: {game}\n"
                                      f"Start Time: {start_time}\n\nThis action cannot be undone!",
                                      parent=pending_window)
        if not response:
            return
        
        csv_filename = get_csv_filename_for_date(date)
        if not os.path.exists(csv_filename):
            messagebox.showerror("Error", f"Record file not found for {date}",
                               parent=pending_window)
            return
        
        try:
            delete_csv_record(csv_filename, date, start_time, game)
            messagebox.showinfo("Success", "Record deleted successfully!", parent=pending_window)
            pending_window.destroy()
            view_pending_payments()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete record: {e}", parent=pending_window)
    
    tk.Button(button_frame, text="Clear Selected Payment", command=clear_selected_pending,
             bg="green", fg="white", width=20, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Delete Selected Record", command=delete_selected_pending,
             bg="darkred", fg="white", width=20, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Refresh",
             command=lambda: [pending_window.destroy(), view_pending_payments()],
             bg="blue", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Close", command=pending_window.destroy,
             bg="gray", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)

def view_statement():
    """View statement for a specific date range or all records"""
    statement_window = tk.Toplevel(root)
    statement_window.title("View Statement")
    statement_window.geometry("1100x650")
    
    date_frame = tk.Frame(statement_window, bg="#e7f3ff", relief=tk.RIDGE, bd=2)
    date_frame.pack(fill="x", padx=10, pady=10)
    
    tk.Label(date_frame, text="Select Date Range:", font=("Arial", 11, "bold"),
            bg="#e7f3ff").pack(side=tk.LEFT, padx=10, pady=8)
    
    tk.Label(date_frame, text="From:", bg="#e7f3ff").pack(side=tk.LEFT, padx=5)
    from_date_entry = tk.Entry(date_frame, width=12, font=("Arial", 10))
    from_date_entry.pack(side=tk.LEFT, padx=5)
    from_date_entry.insert(0, (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
    
    tk.Label(date_frame, text="To:", bg="#e7f3ff").pack(side=tk.LEFT, padx=5)
    to_date_entry = tk.Entry(date_frame, width=12, font=("Arial", 10))
    to_date_entry.pack(side=tk.LEFT, padx=5)
    to_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
    
    tree_frame = tk.Frame(statement_window)
    tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    statement_tree = ttk.Treeview(tree_frame,
                                  columns=("Date", "Customer", "Start", "Game", "Duration",
                                          "Total", "Paid", "Balance", "Status"),
                                  show="headings", yscrollcommand=tree_scroll.set)
    tree_scroll.config(command=statement_tree.yview)
    
    statement_tree.heading("Date", text="Date")
    statement_tree.heading("Customer", text="Customer")
    statement_tree.heading("Start", text="Start Time")
    statement_tree.heading("Game", text="Game")
    statement_tree.heading("Duration", text="Duration")
    statement_tree.heading("Total", text="Total")
    statement_tree.heading("Paid", text="Paid")
    statement_tree.heading("Balance", text="Balance")
    statement_tree.heading("Status", text="Status")
    
    statement_tree.column("Date", width=100)
    statement_tree.column("Customer", width=120)
    statement_tree.column("Start", width=90)
    statement_tree.column("Game", width=100)
    statement_tree.column("Duration", width=80)
    statement_tree.column("Total", width=90, anchor=tk.E)
    statement_tree.column("Paid", width=90, anchor=tk.E)
    statement_tree.column("Balance", width=90, anchor=tk.E)
    statement_tree.column("Status", width=80)
    
    statement_tree.pack(fill="both", expand=True)
    
    summary_frame = tk.Frame(statement_window, bg="#e7f3ff", relief=tk.RIDGE, bd=2)
    summary_frame.pack(fill="x", padx=10, pady=10)
    
    summary_label = tk.Label(summary_frame, text="", font=("Arial", 11, "bold"), bg="#e7f3ff")
    summary_label.pack(pady=10)
    
    def load_statement():
        for item in statement_tree.get_children():
            statement_tree.delete(item)
        
        try:
            from_date = datetime.strptime(from_date_entry.get().strip(), "%Y-%m-%d")
            to_date = datetime.strptime(to_date_entry.get().strip(), "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD",
                               parent=statement_window)
            return
        
        if from_date > to_date:
            messagebox.showerror("Error", "From date must be before To date",
                               parent=statement_window)
            return
        
        records_folder = "records"
        total_amount = 0.0
        paid_amount = 0.0
        pending_amount = 0.0
        record_count = 0
        
        if os.path.exists(records_folder):
            for folder_name in sorted(os.listdir(records_folder)):
                folder_path = os.path.join(records_folder, folder_name)
                if not os.path.isdir(folder_path):
                    continue
                
                try:
                    folder_date = datetime.strptime(folder_name, "%Y-%m-%d")
                    if from_date <= folder_date <= to_date:
                        csv_filename = os.path.join(folder_path, f"{folder_name}.csv")
                        if os.path.exists(csv_filename):
                            with open(csv_filename, "r", newline="", encoding='utf-8') as f:
                                reader = csv.DictReader(f)
                                for row in reader:
                                    date = row.get("Date", "")
                                    customer = row.get("Customer Name", "")
                                    start_time = row.get("Start Time", "")
                                    game = row.get("Game", "")
                                    duration = row.get("Duration", "")
                                    total = row.get("Total Amount", "0")
                                    paid = row.get("Paid Amount", "0")
                                    balance = row.get("Balance Amount", "0")
                                    status = row.get("Payment Status", "PAID")
                                    
                                    statement_tree.insert("", "end",
                                                         values=(date, customer, start_time, game,
                                                                duration, f"Rs. {total}",
                                                                f"Rs. {paid}", f"Rs. {balance}",
                                                                status),
                                                         tags=(status.lower(),))
                                    
                                    try:
                                        total_amount += float(total)
                                        if status == "PAID":
                                            paid_amount += float(total)
                                        else:
                                            pending_amount += float(balance)
                                        record_count += 1
                                    except ValueError:
                                        pass
                except ValueError:
                    continue
        
        statement_tree.tag_configure("paid", background="#d4edda")
        statement_tree.tag_configure("pending", background="#f8d7da")
        
        summary_text = f"Total Records: {record_count} | Total Amount: Rs. {total_amount:.2f} | "
        summary_text += f"Paid: Rs. {paid_amount:.2f} | Pending: Rs. {pending_amount:.2f}"
        summary_label.config(text=summary_text)
    
    button_frame = tk.Frame(statement_window)
    button_frame.pack(fill="x", padx=10, pady=5)
    
    tk.Button(button_frame, text="Load Statement", command=load_statement,
             bg="blue", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Close", command=statement_window.destroy,
             bg="gray", fg="white", width=15, font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
    
    load_statement()

# ==================== SESSION MANAGEMENT ====================
def load_previous_session():
    """Load previous session data including both active and completed sessions"""
    filename = get_recovery_csv_filename()
    today_csv = get_today_csv_filename()
    
    if os.path.exists(filename):
        try:
            with open(filename, "r", newline="", encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if not row.get("End Time", "").strip():
                        date = row.get("Date", datetime.now().strftime("%Y-%m-%d"))
                        customer_name = row.get("Customer Name", "")
                        start_time = row.get("Start Time", "")
                        game = row.get("Game", "")
                        controllers = row.get("Controllers", "2")
                        
                        if start_time and game:
                            tree.insert("", "end",
                                      values=(date, customer_name, start_time, game, "", "", "",
                                             "", "", "", "", controllers, ""),
                                      tags=("active",))
            os.remove(filename)
        except Exception as e:
            print(f"Warning: Could not load recovery file: {e}")
    
    if os.path.exists(today_csv):
        try:
            with open(today_csv, "r", newline="", encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    date = row.get("Date", "")
                    customer_name = row.get("Customer Name", "")
                    start_time = row.get("Start Time", "")
                    game = row.get("Game", "")
                    end_time = row.get("End Time", "")
                    duration = row.get("Duration", "")
                    total = row.get("Total Amount", "")
                    paid = row.get("Paid Amount", "")
                    balance = row.get("Balance Amount", "")
                    cash = row.get("Cash", "")
                    gpay = row.get("GPay", "")
                    controllers = row.get("Controllers", "2")
                    payment_status = row.get("Payment Status", "PAID")
                    
                    if start_time and game:
                        tag = "pending" if payment_status == "PENDING" else "paid" if end_time else "active"
                        tree.insert("", "end",
                                  values=(date, customer_name, start_time, game, end_time,
                                         duration, total, paid, balance, cash, gpay,
                                         controllers, payment_status),
                                  tags=(tag,))
        except Exception as e:
            print(f"Warning: Could not load today's completed sessions: {e}")

def save_all_data_on_exit():
    """Save only active sessions for recovery"""
    filename = get_recovery_csv_filename()
    active_records = []
    
    for item in tree.get_children():
        values = tree.item(item, 'values')
        if len(values) > 4 and not values[4]:
            active_records.append(values)
    
    if active_records:
        try:
            with open(filename, "w", newline="", encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Date", "Customer Name", "Start Time", "Game", "End Time",
                               "Duration", "Total Amount", "Paid Amount", "Balance Amount",
                               "Cash", "GPay", "Controllers", "Payment Status"])
                writer.writerows(active_records)
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save current session data: {e}")
    
    root.destroy()

# ==================== UI HELPER FUNCTIONS ====================
def auto_fill_start_time():
    current_time = datetime.now().strftime("%I:%M %p")
    entry_start.delete(0, tk.END)
    entry_start.insert(0, current_time)

def auto_fill_end_time():
    current_time = datetime.now().strftime("%I:%M %p")
    entry_end.delete(0, tk.END)
    entry_end.insert(0, current_time)
    calculate_duration_preview()

def calculate_duration_preview():
    selected = tree.selection()
    if not selected:
        duration_label.config(text="Duration: Select a customer first")
        return
    
    end_time = entry_end.get().strip()
    if not end_time:
        duration_label.config(text="Duration: Enter end time")
        return
    
    try:
        datetime.strptime(end_time, "%I:%M %p")
    except ValueError:
        duration_label.config(text="Duration: Invalid time format")
        return
    
    values = tree.item(selected[0], "values")
    start_time = values[2]
    game = values[3]
    
    try:
        controllers = int(values[11]) if len(values) > 11 and values[11] else 2
    except (IndexError, ValueError):
        controllers = 2
    
    total, duration_hours = calculate_amount(game, start_time, end_time, controllers)
    
    if total == 0:
        duration_label.config(text="Duration: Invalid time")
        return
    
    entry_total_amount.delete(0, tk.END)
    entry_total_amount.insert(0, f"{total:.2f}")
    
    duration_str = format_duration(duration_hours)
    duration_label.config(text=f"Duration: {duration_str} | Amount: Rs. {total:.2f}")

def on_end_time_change(*args):
    calculate_duration_preview()

def on_total_amount_change(*args):
    """Update the duration label when total amount is manually edited"""
    selected = tree.selection()
    if not selected:
        return
    
    end_time = entry_end.get().strip()
    if not end_time:
        return
    
    try:
        manual_total = float(entry_total_amount.get().strip())
    except ValueError:
        return
    
    values = tree.item(selected[0], "values")
    start_time = values[2]
    game = values[3]
    
    try:
        controllers = int(values[11]) if len(values) > 11 and values[11] else 2
    except (IndexError, ValueError):
        controllers = 2
    
    _, duration_hours = calculate_amount(game, start_time, end_time, controllers)
    
    if duration_hours == 0:
        return
    
    duration_str = format_duration(duration_hours)
    duration_label.config(text=f"Duration: {duration_str} | Amount: Rs. {manual_total:.2f} (Manual)")

def update_live_time():
    """Update the live time display every second"""
    current_time = datetime.now().strftime("%I:%M:%S %p")
    current_date = datetime.now().strftime("%A, %B %d, %Y")
    time_label.config(text=current_time)
    date_label.config(text=current_date)
    root.after(1000, update_live_time)

# ==================== CUSTOMER OPERATIONS ====================
def check_in_customer():
    date = datetime.now().strftime("%Y-%m-%d")
    customer_name = entry_customer_name.get().strip()
    start_time = entry_start.get().strip()
    game = game_var.get().strip()
    
    try:
        controllers = int(controller_var.get())
        if controllers < 1:
            messagebox.showerror("Error", "Controllers must be at least 1")
            return
    except ValueError:
        messagebox.showerror("Error", "Controllers must be a valid number")
        return
    
    if not start_time or not game:
        messagebox.showerror("Error", "Please enter Start Time and select a Game")
        return
    
    try:
        datetime.strptime(start_time, "%I:%M %p")
    except ValueError:
        messagebox.showerror("Error", "Invalid time format. Use HH:MM AM/PM (e.g., 02:30 PM)")
        return
    
    tree.insert("", "end",
               values=(date, customer_name, start_time, game, "", "", "", "", "", "", "",
                      controllers, ""),
               tags=("active",))
    
    entry_customer_name.delete(0, tk.END)
    entry_start.delete(0, tk.END)
    controller_var.set("2")
    game_var.set("")

def check_out_customer():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Warning", "Please select a customer to check-out")
        return
    
    end_time = entry_end.get().strip()
    
    try:
        cash_paid = float(entry_cash.get().strip() or 0)
    except ValueError:
        messagebox.showerror("Error", "Invalid Cash amount entered.")
        return
    
    try:
        gpay_paid = float(entry_gpay.get().strip() or 0)
    except ValueError:
        messagebox.showerror("Error", "Invalid GPay amount entered.")
        return
    
    try:
        manual_total = float(entry_total_amount.get().strip()) if entry_total_amount.get().strip() else None
    except ValueError:
        messagebox.showerror("Error", "Invalid Total Amount entered.")
        return
    
    paid_amount = cash_paid + gpay_paid
    
    if not end_time:
        messagebox.showerror("Error", "Please enter End Time")
        return
    
    try:
        datetime.strptime(end_time, "%I:%M %p")
    except ValueError:
        messagebox.showerror("Error", "Invalid time format. Use HH:MM AM/PM (e.g., 04:30 PM)")
        return
    
    for sel in selected:
        values = list(tree.item(sel, "values"))
        date, customer_name, start_time, game = values[0], values[1], values[2], values[3]
        
        try:
            controllers = int(values[11]) if len(values) > 11 and values[11] else 2
        except Exception:
            controllers = 2
        
        calculated_total, duration_hours = calculate_amount(game, start_time, end_time, controllers)
        
        if calculated_total == 0:
            messagebox.showerror("Error", "Invalid time calculation. End Time must be after Start Time")
            return
        
        total = manual_total if manual_total is not None else calculated_total
        balance_amount = total - paid_amount
        
        if balance_amount > 0.01:
            response = messagebox.askyesno("Payment Pending",
                                          f"Total: Rs. {total:.2f}. Paid: Rs. {paid_amount:.2f}. "
                                          f"Balance: Rs. {balance_amount:.2f}.\n"
                                          f"Mark as PENDING payment?")
            if response:
                payment_status = "PENDING"
            else:
                return
        else:
            payment_status = "PAID"
            balance_amount = 0.0
            customer_name = ""
        
        duration_str = format_duration(duration_hours)
        
        tree.item(sel, values=(date, customer_name, start_time, game, end_time, duration_str,
                              f"{total:.2f}", f"{paid_amount:.2f}", f"{balance_amount:.2f}",
                              f"{cash_paid:.2f}", f"{gpay_paid:.2f}", controllers, payment_status))
        
        if payment_status == "PENDING":
            tree.item(sel, tags=("pending",))
        else:
            tree.item(sel, tags=("paid",))
        
        filename = get_today_csv_filename()
        file_exists = os.path.isfile(filename)
        
        with open(filename, "a", newline="", encoding='utf-8') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["Date", "Customer Name", "Start Time", "Game", "End Time",
                               "Duration", "Total Amount", "Paid Amount", "Balance Amount",
                               "Cash", "GPay", "Controllers", "Payment Status"])
            writer.writerow([date, customer_name, start_time, game, end_time, duration_str,
                           f"{total:.2f}", f"{paid_amount:.2f}", f"{balance_amount:.2f}",
                           f"{cash_paid:.2f}", f"{gpay_paid:.2f}", controllers, payment_status])
        
        status_msg = " (PAYMENT PENDING)" if payment_status == "PENDING" else ""
        amount_msg = (" (Manual Override)" if manual_total is not None and
                     manual_total != calculated_total else "")
        
        messagebox.showinfo("Check-Out Complete",
                          f"Duration: {duration_str}\nTotal: Rs. {total:.2f}{amount_msg}\n"
                          f"Paid: Rs. {paid_amount:.2f}\nBalance: Rs. {balance_amount:.2f}"
                          f"{status_msg}")
    
    entry_end.delete(0, tk.END)
    entry_cash.delete(0, tk.END)
    entry_gpay.delete(0, tk.END)
    entry_total_amount.delete(0, tk.END)
    duration_label.config(text="Duration: --")

def cancel_customer():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Warning", "Please select a customer to cancel")
        return
    
    for sel in selected:
        values = tree.item(sel, "values")
        end_time = values[4] if len(values) > 4 else ""
        
        if end_time:
            messagebox.showwarning("Warning",
                                 "Cannot cancel a completed session. Use Edit if you need to modify it.")
            continue
        
        response = messagebox.askyesno("Confirm Cancel",
                                      "Are you sure you want to remove this customer?\n"
                                      "They will be removed from the list.")
        if response:
            tree.delete(sel)
            messagebox.showinfo("Success", "Customer removed from the list")

def update_payment():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Warning", "Please select a customer to update payment")
        return
    
    try:
        cash_paid = float(entry_cash.get().strip() or 0)
    except ValueError:
        messagebox.showerror("Error", "Invalid Cash amount entered.")
        return
    
    try:
        gpay_paid = float(entry_gpay.get().strip() or 0)
    except ValueError:
        messagebox.showerror("Error", "Invalid GPay amount entered.")
        return
    
    new_paid_amount = cash_paid + gpay_paid
    
    if new_paid_amount <= 0:
        messagebox.showerror("Error", "Please enter a valid payment amount")
        return
    
    for sel in selected:
        values = list(tree.item(sel, "values"))
        
        if len(values) < 13:
            messagebox.showwarning("Warning", "Selected customer record is incomplete or old format.")
            continue
        
        date, customer_name = values[0], values[1]
        start_time, game, end_time, duration = values[2], values[3], values[4], values[5]
        
        try:
            total = float(values[6])
            old_paid = float(values[7])
            old_cash = float(values[9])
            old_gpay = float(values[10])
            controllers = values[11]
            payment_status = values[12]
        except (ValueError, IndexError):
            messagebox.showerror("Error", "Invalid Total/Paid amount in record.")
            continue
        
        if payment_status != "PENDING":
            messagebox.showinfo("Info", "This customer's payment is already marked as PAID")
            continue
        
        final_paid_amount = old_paid + new_paid_amount
        final_balance_amount = total - final_paid_amount
        
        if final_balance_amount > 0.01:
            messagebox.showerror("Error",
                               f"Payment (Rs. {new_paid_amount:.2f}) does not cover the "
                               f"remaining balance. Balance still Rs. {final_balance_amount:.2f}")
            continue
        
        final_payment_status = "PAID"
        final_balance_amount = 0.0
        final_paid_amount = total
        customer_name = ""
        
        final_cash = old_cash + cash_paid
        final_gpay = old_gpay + gpay_paid
        
        tree.item(sel, values=(date, customer_name, start_time, game, end_time, duration,
                              f"{total:.2f}", f"{final_paid_amount:.2f}",
                              f"{final_balance_amount:.2f}", f"{final_cash:.2f}",
                              f"{final_gpay:.2f}", controllers, final_payment_status))
        tree.item(sel, tags=("paid",))
        
        filename = get_csv_filename_for_date(date)
        update_csv_record(filename, date, start_time, game, start_time, game, end_time,
                         duration, f"{total:.2f}", f"{final_paid_amount:.2f}",
                         f"{final_balance_amount:.2f}", f"{final_cash:.2f}",
                         f"{final_gpay:.2f}", controllers, final_payment_status, customer_name)
        
        messagebox.showinfo("Success",
                          f"Payment updated: Rs. {new_paid_amount:.2f} received.\n"
                          f"Total Paid: Rs. {final_paid_amount:.2f}")
    
    entry_cash.delete(0, tk.END)
    entry_gpay.delete(0, tk.END)

# ==================== EDIT OPERATIONS ====================
def open_edit_mode():
    global edit_mode, edit_item_id
    
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Warning", "Please select a customer to edit")
        return
    
    edit_item_id = selected[0]
    values = tree.item(edit_item_id, "values")
    
    if len(values) >= 13:
        date, customer_name, start_time = values[0], values[1], values[2]
        game, end_time, duration = values[3], values[4], values[5]
        total, paid, balance = values[6], values[7], values[8]
        cash, gpay, controllers, payment_status = values[9], values[10], values[11], values[12]
    else:
        date, customer_name, start_time = values[0], values[1], values[2]
        game, end_time, duration = values[3], values[4], values[5]
        total, cash, gpay, controllers = values[6], values[7], values[8], values[9]
        paid = str(float(cash or 0) + float(gpay or 0))
        balance = str(float(total or 0) - float(paid))
        payment_status = values[10] if len(values) > 10 else ""
    
    entry_customer_name.delete(0, tk.END)
    entry_customer_name.insert(0, customer_name)
    
    entry_start.delete(0, tk.END)
    entry_start.insert(0, start_time)
    
    entry_end.delete(0, tk.END)
    entry_end.insert(0, end_time)
    
    entry_total_amount.delete(0, tk.END)
    entry_total_amount.insert(0, total if total else "")
    
    game_var.set(game if game else "")
    controller_var.set(controllers if controllers else "2")
    
    entry_cash.delete(0, tk.END)
    entry_cash.insert(0, cash if cash else "")
    
    entry_gpay.delete(0, tk.END)
    entry_gpay.insert(0, gpay if gpay else "")
    
    edit_mode = True
    btn_check_in.config(text="Save Changes", command=save_edit, bg="purple")
    btn_check_out.config(state="disabled")
    btn_edit.config(state="disabled")
    btn_cancel_edit.config(state="normal")
    btn_cancel_customer.config(state="disabled")
    btn_update_payment.config(state="disabled")
    combo_game.config(state="readonly")

def save_edit():
    global edit_mode, edit_item_id
    
    if not edit_item_id:
        return
    
    values = list(tree.item(edit_item_id, "values"))
    date = values[0]
    customer_name = entry_customer_name.get().strip()
    start_time = entry_start.get().strip()
    game = game_var.get().strip()
    end_time = entry_end.get().strip()
    cash_str = entry_cash.get().strip()
    gpay_str = entry_gpay.get().strip()
    manual_total_str = entry_total_amount.get().strip()
    
    try:
        controllers = int(controller_var.get())
    except ValueError:
        controllers = 2
    
    if not start_time or not game:
        messagebox.showerror("Error", "Start Time and Game are required")
        return
    
    try:
        datetime.strptime(start_time, "%I:%M %p")
    except ValueError:
        messagebox.showerror("Error", "Invalid Start Time format. Use HH:MM AM/PM")
        return
    
    duration = ""
    total = ""
    paid = ""
    balance = ""
    payment_status = ""
    
    if end_time:
        try:
            datetime.strptime(end_time, "%I:%M %p")
        except ValueError:
            messagebox.showerror("Error", "Invalid End Time format. Use HH:MM AM/PM")
            return
        
        calculated_total, duration_hours = calculate_amount(game, start_time, end_time, controllers)
        
        if calculated_total == 0:
            messagebox.showerror("Error",
                               "Invalid time calculation. Make sure End Time is after Start Time.")
            return
        
        try:
            total_amt = float(manual_total_str) if manual_total_str else calculated_total
        except ValueError:
            messagebox.showerror("Error", "Invalid Total Amount entered.")
            return
        
        duration = format_duration(duration_hours)
        
        try:
            paid_amount = float(cash_str or 0) + float(gpay_str or 0)
        except ValueError:
            messagebox.showerror("Error", "Invalid Cash or GPay amount entered.")
            return
        
        balance_amount = total_amt - paid_amount
        
        if balance_amount > 0.01:
            payment_status = "PENDING"
        else:
            payment_status = "PAID"
            balance_amount = 0.0
            customer_name = ""
        
        total = f"{total_amt:.2f}"
        paid = f"{paid_amount:.2f}"
        balance = f"{balance_amount:.2f}"
    
    tree.item(edit_item_id, values=(date, customer_name, start_time, game, end_time, duration,
                                   total, paid, balance, cash_str, gpay_str, controllers,
                                   payment_status))
    
    if payment_status == "PENDING":
        tree.item(edit_item_id, tags=("pending",))
    elif payment_status == "PAID":
        tree.item(edit_item_id, tags=("paid",))
    else:
        tree.item(edit_item_id, tags=("active",))
    
    if end_time:
        filename = get_csv_filename_for_date(date)
        update_csv_record(filename, date, values[2], values[3], start_time, game, end_time,
                         duration, total, paid, balance, cash_str, gpay_str, controllers,
                         payment_status, customer_name)
        messagebox.showinfo("Success", "Customer record updated successfully!")
    else:
        messagebox.showinfo("Success", "Customer record updated (pending check-out)")
    
    cancel_edit()

def cancel_edit():
    global edit_mode, edit_item_id
    
    edit_mode = False
    edit_item_id = None
    
    btn_check_in.config(text="Check-In Customer", command=check_in_customer, bg="blue", fg="white")
    btn_check_out.config(state="normal")
    btn_edit.config(state="normal")
    btn_cancel_edit.config(state="disabled")
    btn_cancel_customer.config(state="normal")
    btn_update_payment.config(state="normal")
    combo_game.config(state="readonly")
    
    entry_customer_name.delete(0, tk.END)
    entry_start.delete(0, tk.END)
    entry_end.delete(0, tk.END)
    entry_cash.delete(0, tk.END)
    entry_gpay.delete(0, tk.END)
    entry_total_amount.delete(0, tk.END)
    controller_var.set("2")
    game_var.set("")
    duration_label.config(text="Duration: --")

# ==================== MAIN APPLICATION ====================
load_game_prices()

root = tk.Tk()
root.title("Gaming Lounge Management System")

win_width = root.winfo_screenwidth()
win_height = root.winfo_screenheight()
root.geometry(f"{win_width}x{win_height}+0+0")

# ==================== TOP CONTAINER ====================
top_container = tk.Frame(root, bg="white")
top_container.pack(side=tk.TOP, fill="x", pady=5)

logo_frame = tk.Frame(top_container, bg="white")
logo_frame.pack(side=tk.LEFT, padx=20)

try:
    logo_image = Image.open("logo.png")
    logo_image = logo_image.resize((150, 80), Image.Resampling.LANCZOS)
    logo_photo = ImageTk.PhotoImage(logo_image)
    logo_label = tk.Label(logo_frame, image=logo_photo, bg="white")
    logo_label.image = logo_photo
    logo_label.pack(pady=5)
    
    business_name = tk.Label(logo_frame, text="Gaming Lounge Management System",
                            font=("Arial", 16, "bold"), bg="white", fg="#2C3E50")
    business_name.pack()
except FileNotFoundError:
    business_name = tk.Label(logo_frame, text="🎮 Gaming Lounge Management System",
                            font=("Arial", 18, "bold"), bg="white", fg="#2C3E50")
    business_name.pack(pady=10)
except Exception as e:
    print(f"Error loading logo: {e}")
    business_name = tk.Label(logo_frame, text="Gaming Lounge Management System",
                            font=("Arial", 16, "bold"), bg="white")
    business_name.pack(pady=10)

time_frame = tk.Frame(top_container, bg="white")
time_frame.pack(side=tk.RIGHT, padx=20)

time_label = tk.Label(time_frame, text="", font=("Arial", 24, "bold"), bg="white", fg="#1a5490")
time_label.pack()

date_label = tk.Label(time_frame, text="", font=("Arial", 12), bg="white", fg="#34495e")
date_label.pack()

# ==================== INPUT FRAME ====================
frame_inputs = tk.Frame(root)
frame_inputs.pack(side=tk.TOP, fill="x", pady=10, padx=10)

tk.Label(frame_inputs, text="Customer Name (Optional):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
entry_customer_name = tk.Entry(frame_inputs, width=15)
entry_customer_name.grid(row=0, column=1, padx=5, pady=5)

tk.Label(frame_inputs, text="Start Time (HH:MM AM/PM):").grid(row=0, column=2, padx=5, pady=5, sticky="e")
entry_start = tk.Entry(frame_inputs, width=15)
entry_start.grid(row=0, column=3, padx=5, pady=5)

btn_auto_start = tk.Button(frame_inputs, text="Auto", command=auto_fill_start_time,
                           bg="lightblue", width=5)
btn_auto_start.grid(row=0, column=4, padx=2, pady=5)

tk.Label(frame_inputs, text="Game:").grid(row=0, column=5, padx=5, pady=5, sticky="e")
game_var = tk.StringVar()
combo_game = ttk.Combobox(frame_inputs, textvariable=game_var,
                         values=sorted(list(GAME_PRICES.keys())),
                         state="readonly", width=15)
combo_game.grid(row=0, column=6, padx=5, pady=5)

tk.Label(frame_inputs, text="End Time (HH:MM AM/PM):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
entry_end = tk.Entry(frame_inputs, width=15)
entry_end.grid(row=1, column=1, padx=5, pady=5)
entry_end.bind('<KeyRelease>', on_end_time_change)

btn_auto_end = tk.Button(frame_inputs, text="Auto", command=auto_fill_end_time,
                         bg="lightgreen", width=5)
btn_auto_end.grid(row=1, column=2, padx=2, pady=5)

tk.Label(frame_inputs, text="Controllers (FIFA/EFOOTBALL):").grid(row=1, column=3, padx=5, pady=5, sticky="e")
controller_var = tk.StringVar(value="2")
entry_controller = tk.Entry(frame_inputs, textvariable=controller_var, width=15)
entry_controller.grid(row=1, column=4, padx=5, pady=5)

tk.Label(frame_inputs, text="Cash:").grid(row=1, column=5, padx=5, pady=5, sticky="e")
entry_cash = tk.Entry(frame_inputs, width=15)
entry_cash.grid(row=1, column=6, padx=5, pady=5)

tk.Label(frame_inputs, text="GPay:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
entry_gpay = tk.Entry(frame_inputs, width=15)
entry_gpay.grid(row=2, column=1, padx=5, pady=5)

tk.Label(frame_inputs, text="Total Amount (Rs):").grid(row=2, column=2, padx=5, pady=5, sticky="e")
entry_total_amount = tk.Entry(frame_inputs, width=15)
entry_total_amount.grid(row=2, column=3, padx=5, pady=5)
entry_total_amount.bind('<KeyRelease>', on_total_amount_change)

tk.Label(frame_inputs, text="(Leave empty for auto-calculation)",
        font=("Arial", 8), fg="gray").grid(row=2, column=4, columnspan=2, padx=5, pady=5, sticky="w")

btn_check_in = tk.Button(frame_inputs, text="Check-In Customer", command=check_in_customer,
                        bg="blue", fg="white", width=15)
btn_check_in.grid(row=3, column=0, columnspan=2, pady=10, padx=5, sticky="ew")

btn_check_out = tk.Button(frame_inputs, text="Check-Out Customer", command=check_out_customer,
                         bg="green", fg="white", width=15)
btn_check_out.grid(row=3, column=2, columnspan=2, pady=10, padx=5, sticky="ew")

btn_cancel_customer = tk.Button(frame_inputs, text="Cancel Customer", command=cancel_customer,
                                bg="red", fg="white", width=13)
btn_cancel_customer.grid(row=3, column=4, pady=10, padx=5, sticky="ew")

btn_update_payment = tk.Button(frame_inputs, text="Update Payment", command=update_payment,
                               bg="darkgreen", fg="white", width=13)
btn_update_payment.grid(row=3, column=5, columnspan=2, pady=10, padx=5, sticky="ew")

btn_edit = tk.Button(frame_inputs, text="Edit Customer", command=open_edit_mode,
                    bg="orange", fg="black", width=12)
btn_edit.grid(row=4, column=0, columnspan=2, pady=5, padx=5, sticky="ew")

btn_cancel_edit = tk.Button(frame_inputs, text="Cancel Edit", command=cancel_edit,
                            bg="gray", fg="white", width=12, state="disabled")
btn_cancel_edit.grid(row=4, column=2, columnspan=2, pady=5, padx=5, sticky="ew")

duration_label = tk.Label(frame_inputs, text="Duration: --", fg="blue", font=("Arial", 10, "bold"))
duration_label.grid(row=5, column=0, columnspan=7, pady=5)

# ==================== BUTTON FRAME ====================
frame_buttons = tk.Frame(root)
frame_buttons.pack(side=tk.TOP, fill="x", pady=5, padx=10)

btn_daily_income = tk.Button(frame_buttons, text="Show Daily Income", command=show_daily_income,
                             bg="gold", fg="black", width=18)
btn_daily_income.pack(side=tk.LEFT, padx=5)

btn_monthly_income = tk.Button(frame_buttons, text="Show Monthly Income", command=show_monthly_income,
                               bg="purple", fg="white", width=18)
btn_monthly_income.pack(side=tk.LEFT, padx=5)

btn_export_excel = tk.Button(frame_buttons, text="Export Financial Report", command=export_to_excel,
                             bg="green", fg="white", width=22)
btn_export_excel.pack(side=tk.LEFT, padx=5)

btn_manual_summary = tk.Button(frame_buttons, text="Add Missing Summary",
                               command=add_manual_daily_summary,
                               bg="#CC0066", fg="white", width=20)
btn_manual_summary.pack(side=tk.LEFT, padx=5)

btn_pending_payments = tk.Button(frame_buttons, text="View Pending Payments",
                                 command=view_pending_payments,
                                 bg="#dc3545", fg="white", width=20)
btn_pending_payments.pack(side=tk.LEFT, padx=5)

btn_view_statement = tk.Button(frame_buttons, text="View Statement", command=view_statement,
                               bg="#17a2b8", fg="white", width=18)
btn_view_statement.pack(side=tk.LEFT, padx=5)

btn_add_expense = tk.Button(frame_buttons, text="Add Expense", command=add_expense,
                            bg="#FF6347", fg="white", width=15)
btn_add_expense.pack(side=tk.LEFT, padx=5)

btn_view_expenses = tk.Button(frame_buttons, text="View Expenses", command=view_expenses,
                              bg="#8B4513", fg="white", width=15)
btn_view_expenses.pack(side=tk.LEFT, padx=5)

btn_manage_games = tk.Button(frame_buttons, text="Manage Games & Prices", command=manage_games,
                             bg="#007BFF", fg="white", width=20, font=("Arial", 10, "bold"))
btn_manage_games.pack(side=tk.LEFT, padx=5)

# ==================== TREE VIEW ====================
tree_frame = tk.Frame(root)
tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

tree_scroll = tk.Scrollbar(tree_frame)
tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

tree = ttk.Treeview(tree_frame,
                   columns=("Date", "Customer Name", "Start", "Game", "End", "Duration",
                           "Total", "Paid", "Balance", "Cash", "GPay", "Controllers", "Status"),
                   show="headings", yscrollcommand=tree_scroll.set)
tree_scroll.config(command=tree.yview)

for col in ("Date", "Customer Name", "Start", "Game", "End", "Duration", "Total", "Paid",
           "Balance", "Cash", "GPay", "Controllers", "Status"):
    tree.heading(col, text=col)
    if col in ("Total", "Paid", "Balance"):
        tree.column(col, width=80, anchor=tk.E)
    elif col in ("Start", "End", "Status", "Cash", "GPay"):
        tree.column(col, width=70)
    elif col in ("Game", "Duration", "Customer Name"):
        tree.column(col, width=90)
    else:
        tree.column(col, width=60)

tree.pack(fill="both", expand=True)

tree.tag_configure("active", background="white")
tree.tag_configure("paid", background="lightgreen")
tree.tag_configure("pending", background="red", foreground="white")

load_previous_session()
update_live_time()

root.protocol("WM_DELETE_WINDOW", save_all_data_on_exit)
root.mainloop()
